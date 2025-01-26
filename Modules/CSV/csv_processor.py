import csv
import os
from openpyxl import load_workbook
from Modules.Transform import data_transformer
from Modules.Utils.file_manager import get_csv_files

def save_list_csv(data: list, path: str):
    """
    Зберігає список у CSV-файл.

    Кожен вкладений список у `data` відповідає одному рядку у CSV-файлі.
    Якщо файл за вказаним шляхом не існує, функція створить його.

    Параметри:
        data (list): Двовимірний список, де кожен підсписок відповідає рядку у CSV.
        path (str): Шлях до файлу, куди зберігатиметься список.

    Помилки:
        - Виводить повідомлення про помилку, якщо шлях до файлу не знайдено (FileNotFoundError).

    Приклад використання:
        save_list_csv([["name", "age"], ["Alice", 30]], "output.csv")
    """
    try:
        with open(path, mode="w", encoding="utf-8", newline="") as csv_file:
            writer = csv.writer(csv_file)
            writer.writerows(data)
    except FileNotFoundError as e:
        print(f'Error: Unable to find file path — {e}')

def open_csv_list(path: str):
    """
    Читає дані з CSV-файлу і повертає їх у вигляді списку.

    Кожен рядок у CSV-файлі перетворюється на список, а всі рядки
    зберігаються у вигляді вкладеного списку.

    Параметри:
        path (str): Шлях до CSV-файлу.

    Повертає:
        list: Вкладений список, де кожен елемент — рядок з файлу.

    Помилки:
        - Виводить повідомлення про помилку, якщо шлях до файлу не знайдено (FileNotFoundError).
        - Якщо виникла помилка, функція повертає порожній список.

    Приклад використання:
        data = open_csv_list("input.csv")
    """
    try:
        with open(path, mode="r", encoding='utf-8') as csv_file:
            data_csv = list(csv.reader(csv_file))
            # print(data_csv)
    except FileNotFoundError as e:
        print(f'Error: Unable to find file path — {e}')
        data_csv = []
    return data_csv

def list_csv_in_creation():
    """
    Збирає всі CSV-файли з папки `CSV_creation` і повертає їх абсолютні шляхи.

    Функція шукає файли у папці `Data_BV/CSV_creation`, яка знаходиться
    відносно кореня проєкту. Використовує абсолютний шлях для пошуку.

    Параметри:
        Немає.

    Повертає:
        list: Список абсолютних шляхів до всіх CSV-файлів у папці.

    Залежності:
        Функція використовує `get_csv_files(directory)` для пошуку файлів.

    Приклад використання:
        csv_files = list_csv_in_creation()
    """
    # Отримуємо абсолютний шлях до папки Data_BV/CSV_creation
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))  # Шлях до кореня проєкту
    directory = os.path.abspath(os.path.join(base_dir, "Data_BV", "CSV_creation"))
    return get_csv_files(directory)

#-----------------To Do-----------------#
def from_csv_data_sort(path: str):
    """
    Завантажує дані з CSV-файлу, сортує їх за першим стовпцем і розділяє на масиви X та Y.

    Функція:
    1. Читає дані з CSV-файлу, використовуючи `open_csv_list`.
    2. Перший рядок (заголовки) зберігає у змінній `top`.
    3. Сортує рядки даних за значенням у першому стовпці (після перетворення на float).
    4. Розділяє дані на два списки:
        - `x`: значення першого стовпця (типу float).
        - `y`: значення другого стовпця (типу float).
    5. Виводить повідомлення про помилки, якщо дані не вдалося перетворити на числа.

    Параметри:
        path (str): Шлях до CSV-файлу.

    Повертає:
        tuple: Кортеж із чотирьох елементів:
            - sorted_data (list): Відсортовані рядки даних.
            - top (list): Заголовки CSV-файлу.
            - x (list): Значення першого стовпця у вигляді чисел (float).
            - y (list): Значення другого стовпця у вигляді чисел (float).

    Помилки:
        - Якщо значення у стовпці неможливо перетворити на число, виводиться повідомлення з індексом і значенням.

    Приклад використання:
        sorted_data, top, x, y = from_csv_data_sort("data.csv")

    Залежності:
        - Використовує функцію `open_csv_list` для завантаження даних із CSV.

    Примітка:
        Функція очікує, що перший стовпець містить значення для сортування (типу float), а другий стовпець містить числові дані.

    """
    data = open_csv_list(path)
    top = data[0]
    data = data[1:]

    sorted_data = sorted(data, key=lambda item: float(item[0].replace(",", ".")))

    x = []
    y = []

    for row in sorted_data:
        for i, value in enumerate(row):
            value = value.replace(",", ".")
            try:
                number = float(value)
                if i == 0:
                    x.append(number)
                elif i == 1:
                    y.append(number)
                else:
                    print(f"Error: [{i}] '{value}' is incorrect data.")
            except ValueError:
                print(f"Error: [{i}] '{value}' is not a valid number.")
    return sorted_data, top, x, y
#-----------------To Do-----------------#


def convert_xlsx_to_csv(xlsx_path: str, output_dir: str, sheet_name: str = "Sheet1"):
    """
    Конвертує дані з XLSX у CSV, зберігаючи пари колонок у окремих CSV-файлах.

    Параметри:
        xlsx_path (str): Шлях до вхідного XLSX-файлу.
        output_dir (str): Директорія, куди будуть збережені CSV-файли.
        sheet_name (str): Назва листа у XLSX-файлі (за замовчуванням "Sheet1").

    Повертає:
        list: Список імен згенерованих CSV-файлів.
    """
    try:
        # Завантажуємо Excel-файл
        workbook = load_workbook(xlsx_path)

        # Перевіряємо, чи існує потрібний лист
        if sheet_name not in workbook.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found in the workbook.")
            return []

        sheet = workbook[sheet_name]

        # Збираємо всі рядки
        data_all_rows = [row for row in sheet.iter_rows(min_row=1, values_only=True)]

        # Перевірка парності колонок
        if not data_all_rows or len(data_all_rows[0]) % 2 != 0:
            print("Error: Odd number of columns in the sheet. Operation stopped.")
            return []

        # Обробка колонок
        csv_files = []
        for column_index in range(0, len(data_all_rows[0]), 2):
            heading = [data_all_rows[0][column_index], data_all_rows[0][column_index + 1]]
            data_to_csv = [heading]

            # Додаємо дані рядків
            for row in data_all_rows[1:]:
                setting = row[column_index]
                kv = row[column_index + 1]
                if setting is not None and kv is not None:
                    data_to_csv.append([setting, kv])

            # Створюємо ім'я файлу
            csv_file_name = f"{heading[0].replace(' ', '_').replace('/', '_')}.csv"
            output_path = os.path.join(output_dir, csv_file_name)

            # Зберігаємо CSV
            save_list_csv(data_to_csv, output_path)
            csv_files.append(csv_file_name)

        return csv_files

    except FileNotFoundError:
        print(f"Error: File '{xlsx_path}' not found.")
        return []
    except Exception as e:
        print(f"Unexpected error: {e}")
        return []


def valve_data_processed(path_csv: str, deg: int):
    """
    Обробляє дані клапана з CSV-файлу, виконує поліноміальне перетворення для даних x та y
    і збирає результат у словник з додатковою інформацією.

    Параметри:
        path_csv (str): Шлях до CSV-файлу з даними клапана.
        deg (int): Ступінь полінома для перетворення даних.

    Повертає:
        dict: Словник з результатами перетворення і додатковими даними клапана
        (name, article_number).
    """
    try:
        # Отримуємо відсортовані дані з CSV
        data = from_csv_data_sort(path_csv)

        # Перевірка наявності необхідних даних
        if len(data) < 4:
            print("Error: Insufficient data in CSV file.")
            return {}

        # Ініціалізація значень
        name = data[1][0]
        art_nr = data[1][1]
        x = data[2]
        y = data[3]

        # Перевірка, чи є дані для x та y
        if not x or not y:
            print("Error: Missing data in x or y values.")
            return {}

        # Поліноміальне перетворення
        valve_data = data_transformer.polynomial_dict(x, y, deg)

        # Додавання інформації про клапан
        valve_data['name'] = name
        valve_data['article_number'] = art_nr

        return valve_data

    except FileNotFoundError:
        print(f"Error: File '{path_csv}' not found.")
        return {}
    except Exception as e:
        print(f"Unexpected error: {e}")
        return {}


def valve_data_aquapresso(path_csv: str, deg: int):
    """
    Обробляє дані клапана з CSV-файлу, виконує поліноміальне перетворення для даних x та y
    і збирає результат у словник з додатковою інформацією for aquapresso.

    Параметри:
        path_csv (str): Шлях до CSV-файлу з даними клапана.
        deg (int): Ступінь полінома для перетворення даних.

    Повертає:
        dict: Словник з результатами перетворення і додатковими даними клапана
        (name, article_number).
    """
    try:
        # Отримуємо відсортовані дані з CSV
        data = from_csv_data_sort(path_csv)

        # Перевірка наявності необхідних даних
        if len(data) < 4:
            print("Error: Insufficient data in CSV file.")
            return {}

        # Ініціалізація значень
        name = data[1][0]
        art_nr = data[1][1]
        x = data[2]
        y = data[3]
        setting_max = max(x)
        setting_min = min(x)
        kv_max = max(y)
        kv_min = min(y)


        # Перевірка, чи є дані для x та y
        if not x or not y:
            print("Error: Missing data in x or y values.")
            return {}

        # Поліноміальне перетворення
        valve_data = data_transformer.polynomial_dict(x, y, deg)

        # Додавання інформації про клапан
        valve_data['name'] = name
        valve_data['article_number'] = art_nr
        valve_data[' setting_max'] = setting_max
        valve_data['setting_min'] = setting_min
        valve_data['kv_max'] = kv_max
        valve_data['kv_min'] = kv_min

        return valve_data

    except FileNotFoundError:
        print(f"Error: File '{path_csv}' not found.")
        return {}
    except Exception as e:
        print(f"Unexpected error: {e}")
        return {}


def prepare_data_for_csv(data: list[dict]) -> list:
    """
    Створює список із даними для запису у форматі CSV.

    Перший рядок містить ключі (заголовки), а наступні рядки — значення
    відповідно до цих ключів. Підготовлений список передається до функції
    `save_list_csv` для запису у файл.

    Параметри:
        data (list[dict]): Список із словників, який потрібно обробити.
        path (str): Шлях до файлу, куди буде записано CSV.

    Повертає:
        list: Список, підготовлений для запису у CSV.

    Помилки:
        - Виводить повідомлення, якщо дані не є списком словників.
    """
    try:
        # Перевірка на те, що дані є списком словників
        if not data or not isinstance(data, list) or not isinstance(data[0], dict):
            print("Error: The provided data must be a list of dictionaries.")
            return []

        # Отримання ключів із першого словника
        keys = list(data[0].keys())
        # Створення нового списку для запису
        csv_data = [keys]  # Перший рядок — це заголовки

        # Додаємо значення кожного словника у список
        for item in data:
            csv_data.append([item.get(key, "") for key in keys])  # Додаємо значення в порядку ключів

        return csv_data

    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return []

# a = valve_data_aquapresso()
# print(a)


# base_dir = os.path.abspath(__file__)
# print(base_dir)
# base_dir = os.path.dirname(os.path.abspath(__file__))
# print(base_dir)
# base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# print(base_dir)
# path_abs = os.path.join(base_dir,'Data_BV', 'CSV_creation','TA-BVS_DN125.csv')
#
# data_v = valve_data_processed(path_abs, 6)
# print(data_v)

####


# xlsx_path1 = 'Data_BV.xlsx'
# output_dir1 = 'CSV_creation'

# print(open_csv_list('CSV_creation/STAD_DN10.csv'))

