import csv
import os
from openpyxl import load_workbook

#Збереження списку у csv
def save_list_csv(data: list, path: str):
    try:
        with open(path, mode="w", encoding="utf-8", newline="") as csv_file:
            writer = csv.writer(csv_file)
            writer.writerows(data)
    except FileNotFoundError as e:
        print(f'Error: Unable to find file path — {e}')

#Закачування даних с csv до списку
def open_csv_list(path: str):
    try:
        with open(path, mode="r", encoding='utf-8') as csv_file:
            data_csv = list(csv.reader(csv_file))
            # print(data_csv)
    except FileNotFoundError as e:
        print(f'Error: Unable to find file path — {e}')
        data_csv = []
    return data_csv

#Збереження даних з xlsx по дві колонки у csv
def convert_xlsx_to_csv(xlsx_path: str, output_dir: str):
    workbook = load_workbook(xlsx_path)
    sheet = workbook["Sheet1"]

    data_all_rows = []

    for row in sheet.iter_rows(min_row=1, values_only=True):
        if len(row)%2 == 0:
            data_all_rows.append(row)
        else:
            print('Odd number of columns, program stopped.')
            break

    column_index = 1

    for x in data_all_rows[0][::2]: #Ітерація по першому рядку з кроком 2
        data_to_csv = []

        heading = [x,data_all_rows[0][column_index]]
        data_to_csv.append(heading)
        csv_file_name = f'{x.replace(' ', '_')}.csv'

        for y in data_all_rows[1:]:
            setting = y[column_index - 1]
            kv = y[column_index]
            if setting is not None and kv is not None:
                data_to_csv.append([setting, kv])
            else:
                pass
        column_index += 2
        output_path = os.path.join(output_dir, csv_file_name)
        save_list_csv(data_to_csv, output_path)

        print(csv_file_name)

# xlsx_path1 = 'Data_BV.xlsx'
# output_dir1 = 'CSV_creation'

# print(open_csv_list('CSV_creation/STAD_DN10.csv'))